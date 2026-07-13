#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""rebuild_data.py v2 (13 ก.ค. 2569) — สร้างชั้นข้อมูลทั้งหมดจากไฟล์ตรวจทาน
กฎ cluster_id: พิกัด <20 ม. = จุดเดียวกัน · ยกเว้นพื้นที่บางบอน (ต้องชื่อตรงกัน)
และประเภทปัญหาต่างกันปี 2567/2568 ไม่รวมเป็นจุดเดียว · 13_Pending เฉพาะจุด กทม.
sol_* คงจาก combined_data.json เดิม (จับคู่ year+point_id)
วิธีใช้: python scripts/rebuild_data.py "<ไฟล์รวม 3 ปี.xlsx>" <path repo>
© Prapawadee_W. · Traffic and Transportation Department, BMA"""
import json, math, re, sys, datetime, unicodedata
import openpyxl

SRC = sys.argv[1]; REPO = sys.argv[2]
SHEETS = {2566: '2566 (266 จุด)', 2567: '2567 (127 จุด)', 2568: '2568 (77 จุด)'}
FIELDS = ['year','point_id','point_name','point_type','district','zone','road','cause',
          'peak_time','agency_main','solution_group','progress_pct',
          'sol_traffic_mgmt','sol_enforcement','sol_signal','sol_physical','sol_busbay','sol_cctv',
          'latitude','longitude','cluster_id']
SOL_KEYS = ['sol_traffic_mgmt','sol_enforcement','sol_signal','sol_physical','sol_busbay','sol_cctv']
NOTE_COLS = {2567: '[เฉพาะปี] หมายเหตุ การดำเนินงาน/ปัญหาอุปสรรค', 2568: '[เฉพาะปี] หมายเหตุ ปัญหาอุปสรรค'}
EXT = ['กรมทางหลวง','กรมทางหลวงชนบท','การทางพิเศษแห่งประเทศไทย','การรถไฟแห่งประเทศไทย']

def fnum(v):
    if v is None or str(v).strip()=='': return None
    return float(v)
def stxt(v):
    if v is None: return None
    s=str(v); return s if s.strip()!='' else None
def nname(s):
    if not s: return ''
    return re.sub(r'[\s\(\)\-\–\.]','',unicodedata.normalize('NFC',str(s)))
def dist_m(a,b,c,d):
    return math.hypot((a-c)*111320,(b-d)*111320*math.cos(math.radians(13.75)))

def load_rows(wb,year):
    ws=wb[SHEETS[year]]
    hdr=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
    out=[]
    for raw in ws.iter_rows(min_row=2,values_only=True):
        if raw[0] is None and raw[1] is None: continue
        d=dict(zip(hdr,raw))
        rec={'year':float(year),'point_id':float(d['รหัสจุด']),
             'point_name':stxt(d['ชื่อจุด/บริเวณ']),'point_type':stxt(d['ประเภทปัญหา']),
             'district':stxt(d['เขต']),'zone':stxt(d['กลุ่มเขต/โซน']),'road':stxt(d['ถนน']),
             'cause':stxt(d['สาเหตุเบื้องต้น']),
             'peak_time':stxt(d.get('[เฉพาะปี] ช่วงเวลาที่การจราจรติดขัด')) if year==2566 else None,
             'agency_main':stxt(d['หน่วยงานรับผิดชอบหลัก']),'solution_group':stxt(d['กลุ่มแนวทางแก้ไข']),
             'progress_pct':fnum(d['ความคืบหน้า/สถานะ']),
             'latitude':float(d['ละติจูด']),'longitude':float(d['ลองจิจูด']),
             '_note':str(d.get(NOTE_COLS.get(year),'') or '').strip()}
        for k in SOL_KEYS: rec[k]=None
        out.append(rec)
    return out

def migrate_sol(raw,repo):
    old=json.load(open(f'{repo}/data/combined_data.json',encoding='utf-8'))
    lut={(int(r['year']),int(r['point_id'])):r for r in old}
    n=0
    for r in raw:
        o=lut.get((int(r['year']),int(r['point_id'])))
        if o and any(o.get(k) is not None for k in SOL_KEYS):
            for k in SOL_KEYS: r[k]=o.get(k)
            n+=1
    return n

def assign_clusters(raw):
    n=len(raw); parent=list(range(n))
    def find(x):
        while parent[x]!=x: parent[x]=parent[parent[x]]; x=parent[x]
        return x
    def union(a,b):
        ra,rb=find(a),find(b)
        if ra!=rb: parent[rb]=ra
    for r in raw:
        r['_n']=nname(r['point_name'])
        r['_bb']=('บางบอน' in str(r['district'] or '')) or ('บางบอน' in str(r['road'] or ''))
        r['_t']=str(r['point_type'] or '').strip() if int(r['year']) in (2567,2568) else None
    for i in range(n):
        for j in range(i+1,n):
            ri,rj=raw[i],raw[j]
            if abs(ri['latitude']-rj['latitude'])>0.0004: continue
            if dist_m(ri['latitude'],ri['longitude'],rj['latitude'],rj['longitude'])>=20: continue
            same=ri['_n']==rj['_n'] and ri['_n']!=''
            if (ri['_bb'] or rj['_bb']) and not same: continue
            if ri['_t'] and rj['_t'] and ri['_t']!=rj['_t'] and not same: continue
            union(i,j)
    roots={}; cid=0
    for i in sorted(range(n),key=lambda i:(raw[i]['year'],raw[i]['point_id'])):
        r=find(i)
        if r not in roots: cid+=1; roots[r]=cid
    for i in range(n): raw[i]['cluster_id']=float(roots[find(i)])
    for r in raw:
        for k in ('_n','_bb','_t'): r.pop(k,None)
    return cid

def classify(note,progress):
    status='ติดปัญหาอุปสรรค' if progress==0 else 'อยู่ระหว่างดำเนินการ'
    if not note and 0<progress<100: note='อยู่ระหว่างดำเนินการ'
    return status,(note or None)

wb=openpyxl.load_workbook(SRC,data_only=True)
raw=[]
for y in (2566,2567,2568): raw+=load_rows(wb,y)
counts={y:sum(1 for r in raw if int(r['year'])==y) for y in (2566,2567,2568)}
assert counts=={2566:266,2567:127,2568:77},counts
n_sol=migrate_sol(raw,REPO)
n_cl=assign_clusters(raw)
rows=[{f:r[f] for f in FIELDS} for r in raw]

for path in ('data/combined_data.json','data/sheets/08_Combined_Data.json'):
    with open(f'{REPO}/{path}','w',encoding='utf-8') as f:
        json.dump(rows,f,ensure_ascii=False,indent=1)

feats=[]
for r in rows:
    props={k:v for k,v in r.items() if k not in ('latitude','longitude')}
    feats.append({'type':'Feature','geometry':{'type':'Point','coordinates':[r['longitude'],r['latitude']]},'properties':props})
with open(f'{REPO}/data/all_points.geojson','w',encoding='utf-8') as f:
    json.dump({'type':'FeatureCollection','features':feats},f,ensure_ascii=False,indent=1)

tracked=[r for r in raw if int(r['year']) in (2567,2568) and r['progress_pct'] is not None
         and str(r['agency_main'] or '').strip() not in EXT]
latest={};years={}
for r in tracked:
    k=r['cluster_id']; years.setdefault(k,set()).add(int(r['year']))
    if k not in latest or (r['year'],r['progress_pct'])>(latest[k]['year'],latest[k]['progress_pct']):
        latest[k]=r
unresolved=sorted([(k,r) for k,r in latest.items() if r['progress_pct']<100],
                  key=lambda t:(t[1]['progress_pct'],-t[1]['year'],t[1]['point_id']))
p13=[]
for i,(k,r) in enumerate(unresolved,start=1):
    grp,txt=classify(r['_note'],r['progress_pct'])
    p13.append({'ลำดับ':i,'ปี':'/'.join(str(x) for x in sorted(years[k])),
                'ชื่อจุด':re.sub(r'\s+',' ',str(r['point_name'])).strip(),
                'เขต':r['district'],'ผู้รับผิดชอบ':r['agency_main'],
                'คืบหน้า (%)':r['progress_pct'],'สถานะ':grp,'หมายเหตุ':txt,
                'lat':r['latitude'],'lon':r['longitude'],'โซน':r['zone']})
with open(f'{REPO}/data/sheets/13_Pending_Update.json','w',encoding='utf-8') as f:
    json.dump(p13,f,ensure_ascii=False,indent=1)

wx=openpyxl.load_workbook(f'{REPO}/data.xlsx')
ws8=wx['08_Combined Data']
hdr8=[ws8.cell(row=3,column=c).value for c in range(1,ws8.max_column+1)]
assert hdr8[:20]==FIELDS[:20],hdr8
ws8.cell(row=3,column=21,value='cluster_id')
if ws8.max_row>3: ws8.delete_rows(4,ws8.max_row-3)
for i,r in enumerate(rows,start=4):
    for j,fld in enumerate(FIELDS,start=1):
        v=r[fld]
        if fld in ('year','point_id','cluster_id') and v is not None: v=int(v)
        ws8.cell(row=i,column=j,value=v)
ws13=wx['13_Pending Update']
ws13.cell(row=1,column=1,value=f'จุดที่ยังดำเนินการไม่แล้วเสร็จ (เฉพาะจุดของ กทม. — จุดนอกอำนาจ กทม. ไม่นำมาคิด) · จุดจริง {len(p13)} จุด')
cols13=['ลำดับ','ปี','ชื่อจุด','เขต','ผู้รับผิดชอบ','คืบหน้า (%)','สถานะ','หมายเหตุ','lat','lon','โซน']
if ws13.max_row>3: ws13.delete_rows(4,ws13.max_row-3)
for j in range(1,ws13.max_column+1): ws13.cell(row=3,column=j,value=None)
for j,c in enumerate(cols13,start=1): ws13.cell(row=3,column=j,value=c)
for i,rec in enumerate(p13,start=4):
    for j,c in enumerate(cols13,start=1): ws13.cell(row=i,column=j,value=rec[c])
wx.save(f'{REPO}/data.xlsx')

man=json.load(open(f'{REPO}/data/manifest.json',encoding='utf-8'))
try:
    import zoneinfo
    now=datetime.datetime.now(zoneinfo.ZoneInfo('Asia/Bangkok'))
except Exception:
    now=datetime.datetime.now()
be=now.replace(year=now.year+543)
ts=be.strftime('%Y-%m-%dT%H:%M:%SZ')
man['generated_at_utc']=ts; man['updated_at']=ts
man['update_note']=f'rebuild จากไฟล์ตรวจทาน 13 ก.ค. 2569 — cluster {n_cl} จุดจริง, จุดค้างดำเนินการ (กทม.) {len(p13)} จุด, sol_* คงเดิม {n_sol} แถว'
with open(f'{REPO}/data/manifest.json','w',encoding='utf-8') as f:
    json.dump(man,f,ensure_ascii=False,indent=2)
print('rows:',counts,'| clusters:',n_cl,'| pending(กทม.):',len(p13),'| sol migrated:',n_sol)
